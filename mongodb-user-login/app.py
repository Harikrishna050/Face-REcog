from flask import Flask, render_template, url_for, request, session, redirect,jsonify,Response,send_from_directory
import cv2
import base64
import requests
import json 
from threading import Thread
from flask_pymongo import PyMongo
from simple_facerec import SimpleFacerec
from genQRcode import GenerateQR
import bcrypt
from openpyxl import load_workbook
from datetime import datetime
import os
import time

# Load the existing Excel workbook
wb = load_workbook("C:/Users/91637/Downloads/attendance/10-04-24.xlsx")
ws = wb.active

# Assuming your existing column headers are "Time stamp" and "Detected Names"
timestamp_column_header = "Time stamp"
column_header = "Name"

app = Flask(__name__)

sfr = SimpleFacerec()
sfr.load_model('C:/Users/91637/Documents/LogIn/mongodb-user-login/trained_model.json')


def generate_qr_async():
    genQR = GenerateQR()
    genQR.QRcodeFunc()


app.config['MONGO_DBNAME'] = 'FaceRecog'
app.config['MONGO_URI'] = 'mongodb+srv://harikrishna050code:hari050@hari.dzogve1.mongodb.net/FaceRecog?retryWrites=true&w=majority'

# Set session lifetime to 1 minutes (600 seconds)
app.config['PERMANENT_SESSION_LIFETIME'] = 60  # 1 minutes in seconds


mongo = PyMongo(app)
# Initialize local storage

@app.route('/')
def index():
    if 'username' in session:  # Check for user in session
        return render_template('layout.html', username=session['username'])
    else:
        return render_template('index.html')

@app.route('/login', methods=['POST'])
def login():
    users = mongo.db.users  
    login_user = users.find_one({'name' : request.form['username']})

    if login_user:
        if bcrypt.checkpw(request.form['pass'].encode('utf-8'), login_user['password']):
            session['username'] = login_user['name']
            return jsonify({"login":"success"})

    return jsonify({"login":"Invalid"})

@app.route('/facultyLogin', methods=['POST'])
def facultyLogin():
    facultyAuth = mongo.db.facultyAuth  
    login_faculty = facultyAuth.find_one({'name' : request.form['username']})

    if login_faculty and bcrypt.checkpw(request.form['pass'].encode('utf-8'), login_faculty['password']):
        session['username'] = login_faculty['name']
        return redirect(url_for('index'))  # Redirect to 'index' route upon successful login
    else:
        error_message = "Invalid credentials. Please try again."
        return render_template('index.html', error_message=error_message)


@app.route('/facultyRegister', methods=['POST', 'GET'])
def facultyRegister():
    if request.method == 'POST':
        facultyAuth = mongo.db.facultyAuth
        existing_faculty = facultyAuth.find_one({'name' : request.form['username']})

        if existing_faculty is None:
            hashpass = bcrypt.hashpw(request.form['pass'].encode('utf-8'), bcrypt.gensalt())
            facultyAuth.insert_one({'name' : request.form['username'], 'password' : hashpass})
            return redirect(url_for('index'))
        
        else:
            error_message = "User Already Exists"
            return render_template('register.html', error_message=error_message)

# #QRcode Gen Request
@app.route('/generate_qr', methods=['GET'])
def generate_qr():
    if request.method == 'GET':
        # Start a new thread to execute generate_qr_async function
        Thread(target=generate_qr_async).start()
        # Render the template
        return render_template('QRPage.html')


@app.route('/register', methods=['POST', 'GET'])
def register():
    if request.method == 'POST':
        users = mongo.db.users
        existing_user = users.find_one({'name' : request.form['username']})

        if existing_user is None:
            hashpass = bcrypt.hashpw(request.form['pass'].encode('utf-8'), bcrypt.gensalt())
            users.insert_one({'name' : request.form['username'], 'password' : hashpass})
            return redirect(url_for('index'))
        
        return 'That username already exists!'

    return render_template('register.html')


@app.route('/logout', methods=['POST'])
def logout():
    session.pop('username')
    return redirect(url_for('index'))


@app.route('/liveliness',methods=['POST'])
def detection():
    print("Received Image")
    file = request.files['image']
    file.save('received.jpg')
    frame = cv2.imread('received.jpg')
    jpg_img = cv2.imencode(".jpg",frame)
    b64_string = base64.b64encode(jpg_img[1]).decode('utf-8')
    url = "https://ping.arya.ai/api/v1/liveness"
    payload = { "doc_base64":b64_string, 
                "req_id":  "str"}
    headers = {
        'token': 'cc20f79aa2316ac0a52ae4e71f84ac4c',
        'content-type':'application/json'
        }
    response = requests.request("POST", url, json=payload, headers=headers)
    parsed_data = json.loads(response.text)
    print(response.text)
    is_real_value = parsed_data['is_real']

    if is_real_value == True:
        return jsonify({'isreal':'true'})
    elif is_real_value == False:
        return jsonify({'isreal':"false"})
    else:
        return jsonify({'isreal': "No response"})
    
@app.route('/detect_faces',methods=['POST'])
def detect_faces():
    print("Received Image")
    file = request.files['image']
    file.save('received.jpg')
    frame = cv2.imread('received.jpg')
    _, face_names = sfr.detect_known_faces(frame)
    return jsonify({'regno':face_names[0]})
  
@app.route('/save_to_excel', methods=['POST'])
def save_to_excel():
    register_number = request.json.get('register_number')
    timestamp_column_index = 1
    column_index = 2
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=1, max_row=1):
        if col[0].value == timestamp_column_header:
                timestamp_column_index = col[0].column
        elif col[0].value == column_header:
                column_index = col[0].column
    if timestamp_column_index and column_index:
        next_row = ws.max_row + 1
        ws.cell(row=next_row, column=timestamp_column_index, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=next_row, column=column_index, value=register_number)
        wb.save("C:/Users/91637/Downloads/attendance/10-04-24.xlsx")
        register_number=""
        return jsonify({'message':'Saved successfully'}) 
    else:
        print("Error: Column headers not found in the spreadsheet.")

@app.route('/download-attendance')
def download_attendance():
    directory = "C:/Users/91637/Downloads/attendance/"
    filename = "10-04-24.xlsx"
    return send_from_directory(directory, filename, as_attachment=True)

if __name__ == '__main__':
    app.secret_key = 'mysecret'
    app.run(debug=True,host='0.0.0.0')